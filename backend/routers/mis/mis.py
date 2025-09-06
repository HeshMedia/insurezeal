from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.security import HTTPBearer
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from typing import List, Optional
from config import get_db
from routers.auth.auth import get_current_user
from dependencies.rbac import (
    require_admin_read,
    require_admin_write,
    require_permission
)
from .schemas import (
    MasterSheetResponse,
    MasterSheetRecord,
    BulkUpdateRequest,
    BulkUpdateResponse,
    MasterSheetStatsResponse,
    AgentMISResponse,
    AgentMISRecord,
    AgentMISStats,
    QuarterlySheetUpdateRequest,
    QuarterlySheetUpdateResponse,
    PolicyDocumentsResponse
)
from .helpers import MISHelpers
from typing import Optional, Dict, Any
import logging
import csv
import io
import tempfile
import zipfile
import zipfile
import tempfile

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/mis", tags=["MIS - Management Information System"])
security = HTTPBearer()

mis_helpers = MISHelpers()


@router.get("/quarter-sheet", response_model=MasterSheetResponse)
async def get_quarter_sheet_data(
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(50, ge=1, le=1000, description="Items per page"),
    quarter: Optional[int] = Query(None, ge=1, le=4, description="Quarter number (1-4) to fetch data from specific quarterly sheet"),
    year: Optional[int] = Query(None, ge=2020, le=2030, description="Year to fetch data from specific quarterly sheet"),
    search: Optional[str] = Query(None, description="Search across key fields"),
    
    # Multiple filtering options - support arrays
    agent_code: Optional[List[str]] = Query(None, description="Filter by agent codes (can specify multiple)"),
    insurer_name: Optional[List[str]] = Query(None, description="Filter by insurer names (can specify multiple)"),
    policy_number: Optional[List[str]] = Query(None, description="Filter by policy numbers (can specify multiple)"),
    reporting_month: Optional[List[str]] = Query(None, description="Filter by reporting months (can specify multiple)"),
    child_id: Optional[List[str]] = Query(None, description="Filter by child IDs (can specify multiple)"),
    broker_name: Optional[List[str]] = Query(None, description="Filter by broker names (can specify multiple)"),
    product_type: Optional[List[str]] = Query(None, description="Filter by product types (can specify multiple)"),
    plan_type: Optional[List[str]] = Query(None, description="Filter by plan types (can specify multiple)"),
    make_model: Optional[List[str]] = Query(None, description="Filter by make/models (can specify multiple)"),
    model: Optional[List[str]] = Query(None, description="Filter by models (can specify multiple)"),
    gvw: Optional[List[str]] = Query(None, description="Filter by GVW values (can specify multiple)"),
    rto: Optional[List[str]] = Query(None, description="Filter by RTO codes (can specify multiple)"),
    state: Optional[List[str]] = Query(None, description="Filter by states (can specify multiple)"),
    fuel_type: Optional[List[str]] = Query(None, description="Filter by fuel types (can specify multiple)"),
    cc: Optional[List[str]] = Query(None, description="Filter by CC values (can specify multiple)"),
    age_year: Optional[List[str]] = Query(None, description="Filter by age in years (can specify multiple)"),
    
    # Sorting options
    sort_by: Optional[str] = Query(None, description="Field to sort by (e.g., reporting_month, agent_code, insurer_name, etc.)"),
    sort_order: Optional[str] = Query("asc", regex="^(asc|desc)$", description="Sort order: 'asc' for ascending, 'desc' for descending"),
    
    current_user = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    _rbac_check = Depends(require_admin_read)
):
    """
    Get paginated data from Master Google Sheet or specific Quarterly Sheet
    
    **Admin/SuperAdmin only endpoint**
    
    This endpoint fetches data from either:
    - Master Google Sheet (when quarter/year not specified) - single source of truth for all data
    - Specific Quarterly Sheet (when quarter and year specified) - Q{quarter}-{year} format
    
    **Features:**
    - Paginated results for large datasets
    - Search across multiple key fields
    - Multiple value filtering (arrays supported)
    - Sorting by any field with ascending/descending order
    - All sheet columns included
    - Quarterly sheet targeting for specific periods
    
    **Parameters:**
    - **quarter**: Optional quarter number (1-4). If provided, year must also be specified
    - **year**: Optional year (2020-2030). If provided, quarter must also be specified
    - When both quarter and year are provided, data is fetched from Q{quarter}-{year} sheet
    - When neither is provided, data is fetched from Master sheet
    
    **Search & Filtering:**
    - **search**: Searches across policy number, agent code, customer name, insurer name, broker name, registration number
    - **Multiple Filter Support**: All filter fields support multiple values (use multiple query parameters)
      - **agent_code**: Filter by agent codes (can specify multiple: ?agent_code=A001&agent_code=A002)
      - **insurer_name**: Filter by insurer names (can specify multiple)
      - **policy_number**: Filter by policy numbers (can specify multiple)
      - **reporting_month**: Filter by reporting months (can specify multiple)
      - **child_id**: Filter by child IDs (can specify multiple)
      - **broker_name**: Filter by broker names (can specify multiple)
      - **product_type**: Filter by product types (can specify multiple)
      - **plan_type**: Filter by plan types (can specify multiple)
      - **make_model**: Filter by make/models (can specify multiple)
      - **model**: Filter by models (can specify multiple)
      - **gvw**: Filter by GVW values (can specify multiple)
      - **rto**: Filter by RTO codes (can specify multiple)
      - **state**: Filter by states (can specify multiple)
      - **fuel_type**: Filter by fuel types (can specify multiple)
      - **cc**: Filter by CC values (can specify multiple)
      - **age_year**: Filter by age in years (can specify multiple)
    
    **Sorting:**
    - **sort_by**: Field to sort by (any field from the data)
    - **sort_order**: Sort order - 'asc' for ascending (default), 'desc' for descending
    
    **Returns:**
    - Complete record data with all sheet fields
    - Pagination metadata
    - Row numbers for update operations
    - Sheet name being accessed
    """
    
    try:
        # Validate quarter and year parameters
        if (quarter is not None and year is None) or (quarter is None and year is not None):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Both quarter and year must be provided together, or neither should be provided"
            )
        
        # Determine sheet name and data source
        if quarter is not None and year is not None:
            sheet_name = f"Q{quarter}-{year}"
            data_source = f"quarterly sheet '{sheet_name}'"
        else:
            sheet_name = "Master"
            data_source = "Master sheet"
        
        # Build filters dictionary - handle multiple values
        filters = {}
        
        # Helper function to process list filters
        def add_filter(filter_name: str, filter_values: Optional[List[str]]):
            if filter_values:
                # Remove None values and empty strings
                clean_values = [v for v in filter_values if v and v.strip()]
                if clean_values:
                    filters[filter_name] = clean_values
        
        # Process all filters
        add_filter('agent_code', agent_code)
        add_filter('insurer_name', insurer_name)
        add_filter('policy_number', policy_number)
        add_filter('reporting_month', reporting_month)
        add_filter('child_id', child_id)
        add_filter('broker_name', broker_name)
        add_filter('product_type', product_type)
        add_filter('plan_type', plan_type)
        add_filter('make_model', make_model)
        add_filter('model', model)
        add_filter('gvw', gvw)
        add_filter('rto', rto)
        add_filter('state', state)
        add_filter('fuel_type', fuel_type)
        add_filter('cc', cc)
        add_filter('age_year', age_year)
        
        # Build sorting parameters
        sort_params = {}
        if sort_by:
            sort_params['sort_by'] = sort_by
            sort_params['sort_order'] = sort_order or 'asc'
        
        logger.info(f"Fetching {data_source} data - Page: {page}, Size: {page_size}, Search: '{search}', Filters: {filters}, Sort: {sort_params}")
        
        # Get data from appropriate sheet
        if quarter is not None and year is not None:
            # Fetch from specific quarterly sheet
            result = await mis_helpers.get_quarterly_sheet_data(
                quarter=quarter,
                year=year,
                page=page,
                page_size=page_size,
                search=search,
                filter_by=filters if filters else None,
                sort_by=sort_params.get('sort_by'),
                sort_order=sort_params.get('sort_order', 'asc')
            )
            
            logger.info(f"Successfully retrieved {len(result.records)} records from {data_source}")
            logger.info(f"Returning data from {data_source}")
            
            # Return quarterly sheet response directly
            return result
        else:
            # Fetch from master sheet (existing logic)
            result = await mis_helpers.get_master_sheet_data(
                page=page,
                page_size=page_size,
                search=search,
                filter_by=filters if filters else None,
                sort_by=sort_params.get('sort_by'),
                sort_order=sort_params.get('sort_order', 'asc')
            )
            
            if "error" in result:
                raise HTTPException(
                    status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                    detail=f"{data_source.capitalize()} access error: {result['error']}"
                )
            
            logger.info(f"Successfully retrieved {len(result['records'])} records from {data_source}")
            
            # Add sheet name to response for clarity
            response_data = MasterSheetResponse(
                records=result["records"],
                total_count=result["total_count"],
                page=result["page"],
                page_size=result["page_size"],
                total_pages=result["total_pages"]
            )
            
            # Add sheet information to the response (if the schema supports it)
            logger.info(f"Returning data from {data_source}")
            return response_data
        
    except HTTPException:
        raise
    except Exception as e:
        data_source_text = f"quarterly sheet Q{quarter}-{year}" if quarter and year else "master sheet"
        logger.error(f"Error in get_master_sheet_data for {data_source_text}: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch {data_source_text} data"
        )


@router.put("/quarterly-sheet/update", response_model=QuarterlySheetUpdateResponse)
async def update_quarterly_sheet_records(
    update_request: QuarterlySheetUpdateRequest,
    quarter: int = Query(..., ge=1, le=4, description="Quarter number (1-4)"),
    year: int = Query(..., ge=2020, le=2030, description="Year"),
    current_user = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    _rbac_check = Depends(require_admin_write)
):
    """
    Update quarterly sheet records using direct field mapping
    
    **Admin/SuperAdmin only endpoint**
    
    This endpoint allows updating quarterly sheet records using the actual field names
    as keys in the request body, providing a more intuitive API for quarterly sheet updates.
    Records are identified by their policy number.
    
    **Features:**
    - Direct field mapping using quarterly sheet headers as keys
    - Update multiple records in one request
    - Records identified by policy number
    - All quarterly sheet fields supported
    - Automatic field name to header mapping
    - Detailed success/failure tracking
    
    **Request Format:**
    ```json
    {
      "records": [
        {
          "policy_number": "POL123456789",
          "Reporting Month (mmm'yy)": "Sep'25",
          "Child ID/ User ID [Provided by Insure Zeal]": "IZ001234",
          "Agent Code": "AG001",
          "Customer Name": "John Doe",
          "Broker Name": "ABC Insurance",
          "Insurer name": "XYZ Insurance",
          "Gross premium": "15000",
          "Net premium": "12711",
          "Running Bal": "1016",
          "Agent Total PO Amount": "2500",
          "Actual Agent_PO%": "15.5",
          "Match": "TRUE"
        }
      ]
    }
    ```
    
    **Field Names:**
    Use the exact quarterly sheet header names. All 70+ fields are supported including:
    - Basic Info: "Agent Code", "Policy number", "Customer Name"
    - Insurance: "Broker Name", "Insurer name", "Product Type"
    - Financial: "Gross premium", "Net premium", "Running Bal", "Agent Total PO Amount"
    - Agent Performance: "Actual Agent_PO%", "Agent Commission %"
    - Vehicle: "Registration.no", "Make_Model", "Fuel Type"
    - Status: "Invoice Status", "Remarks"
    
    **Restricted Fields:**
    - "Match" field is read-only and cannot be updated via API
    - The Match status is managed by the system based on data validation rules
    
    **Note:** The policy_number field is required to identify which record to update.
    It should match the "Policy number" value in the quarterly sheet.
    
    **New Fields Support:**
    The endpoint now supports the newly added quarterly sheet fields:
    - "Agent Total PO Amount": Total agent policy office amount
    - "Actual Agent_PO%": Actual agent policy office percentage
    
    **Returns:**
    - Update success/failure summary
    - Processing time
    - Error details for failed updates
    """
    
    try:
        admin_user_id = current_user["user_id"]
        sheet_name = f"Q{quarter}-{year}"
        
        logger.info(f"Starting quarterly sheet update for {len(update_request.records)} records by admin {admin_user_id} on {sheet_name}")
        
        # Convert the new format to the existing bulk update format
        bulk_updates = []
        blocked_fields = []
        
        for record in update_request.records:
            policy_number = record.policy_number
            
            # Get all field updates for this record (excluding policy_number)
            record_dict = record.dict(by_alias=True, exclude_unset=True)
            record_dict.pop('policy_number', None)  # Remove policy_number from updates
            
            # Convert each field to bulk update format
            for field_name, new_value in record_dict.items():
                # Block updates to Match field (case-insensitive check)
                if field_name.lower().strip() in ['match', 'match status']:
                    blocked_fields.append(f"Field '{field_name}' is read-only and cannot be updated")
                    continue
                
                if new_value is not None:  # Only update fields that have values
                    bulk_updates.append({
                        "record_id": policy_number,  # Use policy_number as record_id for bulk update
                        "field_name": field_name,
                        "new_value": str(new_value) if new_value is not None else ""
                    })
        
        # If there are blocked fields, return error
        if blocked_fields:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Update blocked: {'; '.join(blocked_fields)}. The Match field is read-only and managed by the system."
            )
        
        if not bulk_updates:
            return QuarterlySheetUpdateResponse(
                message=f"No fields to update in {sheet_name}",
                total_records=len(update_request.records),
                successful_updates=0,
                failed_updates=0,
                processing_time_seconds=0.0
            )
        
        logger.info(f"Converted {len(update_request.records)} records to {len(bulk_updates)} field updates")
        
        # Use existing bulk update helper with converted data
        from .schemas import BulkUpdateField
        bulk_update_fields = [BulkUpdateField(**update) for update in bulk_updates]
        
        result = await mis_helpers.bulk_update_quarterly_sheet(
            updates=bulk_update_fields,
            quarter=quarter,
            year=year,
            admin_user_id=admin_user_id
        )
        
        logger.info(f"Quarterly sheet update completed on {sheet_name}: {result['successful_updates']} successful, {result['failed_updates']} failed")
        
        return QuarterlySheetUpdateResponse(
            message=f"Quarterly sheet {sheet_name} update completed: {result['successful_updates']} successful, {result['failed_updates']} failed",
            total_records=len(update_request.records),
            successful_updates=result["successful_updates"],
            failed_updates=result["failed_updates"],
            processing_time_seconds=result["processing_time_seconds"]
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in update_quarterly_sheet_records for Q{quarter}-{year}: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update quarterly sheet Q{quarter}-{year}: {str(e)}"
        )

@router.get("/master-sheet/stats", response_model=MasterSheetStatsResponse)
async def get_master_sheet_statistics(
    current_user = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    _rbac_check = Depends(require_admin_read)
):
    """
    Get complete data from Summary Google Sheet
    
    **Admin/SuperAdmin only endpoint**
    
    This endpoint returns the entire contents of the Summary sheet from Google Sheets,
    which contains calculated agent summaries, financial data, and performance metrics.
    
    **Returns:**
    - Complete Summary sheet data as structured JSON
    - All columns and rows from the Summary sheet
    - Headers and data in easy-to-process format
    - Real-time data directly from Google Sheets
    
    **Data Includes:**
    - Agent financial summaries
    - Performance calculations
    - Commission details
    - Balance information
    - All other calculated fields from Summary sheet
    
    **Use Cases:**
    - Dashboard overview widgets
    - Complete financial reporting
    - Agent performance analysis
    - Summary data export
    """
    
    try:
        logger.info("Generating master sheet statistics")
        
        # Get statistics from master sheet
        stats = await mis_helpers.get_master_sheet_stats()
        
        if "error" in stats:
            raise HTTPException(
                status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                detail=f"Master sheet statistics error: {stats['error']}"
            )
        
        logger.info("Successfully generated master sheet statistics")
        
        return MasterSheetStatsResponse(**stats)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in get_master_sheet_statistics: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to generate master sheet statistics"
        )


@router.get("/master-sheet/fields")
async def get_master_sheet_fields(
    quarter: Optional[int] = Query(None, ge=1, le=4, description="Quarter number (1-4) to get quarterly sheet fields"),
    year: Optional[int] = Query(None, ge=2020, le=2030, description="Year to get quarterly sheet fields"),
    current_user = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    _rbac_check = Depends(require_admin_read)
):
    """
    Get list of all available fields in Quarterly Sheets or specific Quarterly Sheet
    
    **Admin/SuperAdmin only endpoint**
    
    Returns the complete list of field names (headers) available in quarterly sheets.
    This is useful for building dynamic UIs and validating field names for updates.
    
    **Parameters:**
    - **quarter**: Optional quarter number (1-4). If provided, year must also be specified
    - **year**: Optional year (2020-2030). If provided, quarter must also be specified
    - When both quarter and year are provided, fields are fetched from Q{quarter}-{year} sheet
    - When neither is provided, fields are fetched from current quarterly sheet standard headers
    
    **Returns:**
    - List of all quarterly sheet column headers (70+ fields)
    - Field descriptions where available
    - Data type hints for each field
    - Sheet name being queried
    
    **Note:** This route now returns quarterly sheet headers by default instead of master sheet headers,
    as quarterly sheets are the current standard for data management.
    """
    
    try:
        # Validate quarter and year parameters
        if (quarter is not None and year is None) or (quarter is None and year is not None):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Both quarter and year must be provided together, or neither should be provided"
            )
        
        if quarter is not None and year is not None:
            # Get quarterly sheet fields
            sheet_name = f"Q{quarter}-{year}"
            from utils.quarterly_sheets_manager import quarterly_manager
            
            quarterly_sheet = quarterly_manager.get_quarterly_sheet(quarter, year)
            if not quarterly_sheet:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"Quarterly sheet {sheet_name} not found"
                )
            
            # Get headers from the actual quarterly sheet
            headers = quarterly_sheet.row_values(1)
            sheet_type = "Quarterly Sheet"
            
            logger.info(f"Retrieved {len(headers)} headers from quarterly sheet {sheet_name}")
            
        else:
            # Get quarterly sheet headers from quarterly manager (current standard headers)
            from utils.quarterly_sheets_manager import quarterly_manager
            headers = quarterly_manager.create_quarterly_sheet_headers()
            sheet_name = "Current Quarterly Sheet Standard"
            sheet_type = "Quarterly Sheet Standard Headers"
            
            logger.info(f"Retrieved {len(headers)} standard quarterly sheet headers")
        
        # Add field descriptions and types
        field_info = []
        for header in headers:
            field_info.append({
                "field_name": header,
                "description": f"{sheet_type} field: {header}",
                "data_type": "string",  # Most fields are strings in sheets
                "updatable": True
            })
        
        return {
            "sheet_name": sheet_name,
            "sheet_type": sheet_type,
            "total_fields": len(headers),
            "fields": field_info,
            "note": f"Use exact field names for bulk update operations on {sheet_name}"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        sheet_text = f"quarterly sheet Q{quarter}-{year}" if quarter and year else "master sheet"
        logger.error(f"Error getting {sheet_text} fields: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to get {sheet_text} fields"
        )


@router.get("/quarterly-sheet/export")
async def export_quarterly_sheet_data(
    quarters: str = Query(..., description="Comma-separated quarters (e.g., '1,2' or '1'). Each quarter requires corresponding year."),
    years: str = Query(..., description="Comma-separated years (e.g., '2025,2025' or '2025'). Must match the number of quarters."),
    format: str = Query("csv", description="Export format: csv, xlsx, or json"),
    search: Optional[str] = Query(None, description="Filter data before export"),
    agent_code: Optional[str] = Query(None, description="Filter by agent code"),
    current_user = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    _rbac_check = Depends(require_admin_read)
):
    """
    Export Quarterly Google Sheet data with Summary sheets in various formats
    
    **Admin/SuperAdmin only endpoint**
    
    Export quarterly sheet data along with corresponding summary sheets for specified quarters.
    Supports multiple quarters and returns both quarterly data and summary calculations.
    
    **Parameters:**
    - **quarters**: Comma-separated quarter numbers (1-4) e.g., "1,2" or "1"
    - **years**: Comma-separated years (2020-2030) e.g., "2025,2025" or "2025"
    - **format**: Export format - csv, xlsx, or json
    - **search**: Filter data before export
    - **agent_code**: Filter by specific agent code
    
    **Formats:**
    - **csv**: Returns ZIP file with separate CSV files for each quarter's data and summary
    - **xlsx**: Returns Excel file with multiple sheets (quarterly data + summary for each quarter)
    - **json**: Returns JSON with quarterly data and summary data for all quarters
    
    **Returns:**
    For each quarter, two datasets are included:
    1. **Quarterly Sheet Data**: Complete quarterly sheet with all fields and records (including new fields: "Agent Total PO Amount", "Actual Agent_PO%")
    2. **Summary Sheet Data**: Calculated summaries, running balances, and agent statistics
    
    **Examples:**
    - Single quarter: quarters=1&years=2025 (Q1-2025 data + summary)
    - Multiple quarters: quarters=1,2&years=2025,2025 (Q1-2025 and Q2-2025 data + summaries)
    - Cross-year: quarters=4,1&years=2024,2025 (Q4-2024 and Q1-2025 data + summaries)
    
    **Note:** Large datasets are automatically handled. XLSX format recommended for multiple quarters.
    All exports include the complete 70+ field quarterly sheet structure with newly added fields.
    """
    
    try:
        # Parse and validate quarters and years
        quarter_list = [int(q.strip()) for q in quarters.split(',')]
        year_list = [int(y.strip()) for y in years.split(',')]
        
        if len(quarter_list) != len(year_list):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Number of quarters must match number of years"
            )
        
        # Validate quarter and year ranges
        for quarter in quarter_list:
            if quarter < 1 or quarter > 4:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Invalid quarter: {quarter}. Must be between 1 and 4"
                )
        
        for year in year_list:
            if year < 2020 or year > 2030:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Invalid year: {year}. Must be between 2020 and 2030"
                )
        
        if format not in ["csv", "xlsx", "json"]:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid format. Supported formats: csv, xlsx, json"
            )
        
        logger.info(f"Exporting quarterly data for quarters: {quarter_list}, years: {year_list}, format: {format}")
        
        # Collect data for all requested quarters
        all_quarterly_data = {}
        all_summary_data = {}
        
        filters = {}
        if agent_code:
            filters['agent_code'] = agent_code
        
        for quarter, year in zip(quarter_list, year_list):
            sheet_name = f"Q{quarter}-{year}"
            
            try:
                # Get quarterly sheet data
                quarterly_result = await mis_helpers.get_quarterly_sheet_data(
                    quarter=quarter,
                    year=year,
                    page=1,
                    page_size=1000000,  # Large number to get all records
                    search=search,
                    filter_by=filters if filters else None
                )
                
                # Convert records to dict format for export
                quarterly_records = []
                if hasattr(quarterly_result, 'records') and quarterly_result.records:
                    for record in quarterly_result.records:
                        if hasattr(record, 'dict'):
                            # Pydantic model - convert to dict
                            quarterly_records.append(record.dict(by_alias=True))
                        elif isinstance(record, dict):
                            # Already a dict
                            quarterly_records.append(record)
                        else:
                            # Try to convert to dict using vars()
                            quarterly_records.append(vars(record) if hasattr(record, '__dict__') else str(record))
                elif isinstance(quarterly_result, dict) and quarterly_result.get("records"):
                    quarterly_records = quarterly_result["records"]
                
                # Get summary data for this quarter
                try:
                    summary_result = await mis_helpers.get_quarterly_summary_data(
                        quarter=quarter,
                        year=year,
                        agent_code=agent_code
                    )
                except AttributeError:
                    # Fallback if method doesn't exist - create basic summary from quarterly data
                    logger.info(f"get_quarterly_summary_data not available, creating basic summary from quarterly data")
                    summary_result = []
                    
                    # Create basic summary from quarterly records if we have data
                    if quarterly_records:
                        # Calculate basic statistics for the quarter
                        total_policies = len(quarterly_records)
                        total_gross_premium = 0.0
                        total_net_premium = 0.0
                        total_running_balance = 0.0
                        total_commissionable_premium = 0.0
                        
                        agents_summary = {}
                        
                        for record in quarterly_records:
                            # Skip if agent_code filter is applied and doesn't match
                            record_agent_code = record.get("Agent Code", "")
                            if agent_code and record_agent_code != agent_code:
                                continue
                                
                            # Initialize agent summary if not exists
                            if record_agent_code not in agents_summary:
                                agents_summary[record_agent_code] = {
                                    "Agent Code": record_agent_code,
                                    "Total Policies": 0,
                                    "Total Gross Premium": 0.0,
                                    "Total Net Premium": 0.0,
                                    "Total Running Balance": 0.0,
                                    "Total Commissionable Premium": 0.0,
                                    "Quarter": f"Q{quarter}-{year}"
                                }
                            
                            # Add to agent summary
                            agent_summary = agents_summary[record_agent_code]
                            agent_summary["Total Policies"] += 1
                            
                            # Safely convert and add financial values
                            try:
                                gross_premium = float(str(record.get("Gross premium", "0")).replace(",", "").replace("₹", "")) if record.get("Gross premium") else 0.0
                                net_premium = float(str(record.get("Net premium", "0")).replace(",", "").replace("₹", "")) if record.get("Net premium") else 0.0
                                running_balance = float(str(record.get("Running Bal", "0")).replace(",", "").replace("₹", "")) if record.get("Running Bal") else 0.0
                                commissionable_premium = float(str(record.get("Commissionable Premium", "0")).replace(",", "").replace("₹", "")) if record.get("Commissionable Premium") else 0.0
                                
                                agent_summary["Total Gross Premium"] += gross_premium
                                agent_summary["Total Net Premium"] += net_premium
                                agent_summary["Total Running Balance"] += running_balance
                                agent_summary["Total Commissionable Premium"] += commissionable_premium
                                
                                total_gross_premium += gross_premium
                                total_net_premium += net_premium
                                total_running_balance += running_balance
                                total_commissionable_premium += commissionable_premium
                                
                            except (ValueError, TypeError) as e:
                                logger.debug(f"Error converting financial values for record: {e}")
                                continue
                        
                        # Convert agents summary to list
                        summary_result = list(agents_summary.values())
                        
                        # Add overall summary if no agent filter
                        if not agent_code and summary_result:
                            summary_result.append({
                                "Agent Code": "TOTAL",
                                "Total Policies": total_policies,
                                "Total Gross Premium": total_gross_premium,
                                "Total Net Premium": total_net_premium,
                                "Total Running Balance": total_running_balance,
                                "Total Commissionable Premium": total_commissionable_premium,
                                "Quarter": f"Q{quarter}-{year}"
                            })
                        
                        logger.info(f"Created basic summary for {len(summary_result)} agents in {sheet_name}")
                    else:
                        logger.info(f"No quarterly records found to create summary for {sheet_name}")
                        summary_result = []
                
                all_quarterly_data[sheet_name] = quarterly_records
                all_summary_data[f"{sheet_name}_Summary"] = summary_result if summary_result else []
                
                logger.info(f"Retrieved {len(quarterly_records)} records from {sheet_name}")
                
            except Exception as e:
                logger.warning(f"Could not retrieve data for {sheet_name}: {str(e)}")
                all_quarterly_data[sheet_name] = []
                all_summary_data[f"{sheet_name}_Summary"] = []
        
        # Handle different export formats
        if format == "json":
            return JSONResponse(content={
                "quarterly_data": all_quarterly_data,
                "summary_data": all_summary_data,
                "export_info": {
                    "quarters_requested": quarter_list,
                    "years_requested": year_list,
                    "total_sheets": len(quarter_list) * 2,  # quarterly + summary for each
                    "filters_applied": filters,
                    "search_term": search
                }
            }, headers={
                "Content-Disposition": f"attachment; filename=quarterly_export_Q{'-'.join(map(str, quarter_list))}_{'-'.join(map(str, year_list))}.json"
            })
        
        elif format == "csv":
            import zipfile
            import tempfile
            
            # Create a temporary zip file with CSV files for each sheet
            temp_dir = tempfile.mkdtemp()
            zip_path = f"{temp_dir}/quarterly_export.zip"
            
            with zipfile.ZipFile(zip_path, 'w') as zip_file:
                # Add quarterly data CSV files
                for sheet_name, records in all_quarterly_data.items():
                    if records:
                        csv_content = io.StringIO()
                        headers = list(records[0].keys()) if isinstance(records[0], dict) else [f"col_{i}" for i in range(len(records[0]))]
                        writer = csv.DictWriter(csv_content, fieldnames=headers)
                        writer.writeheader()
                        for record in records:
                            writer.writerow(record if isinstance(record, dict) else dict(zip(headers, record)))
                        
                        zip_file.writestr(f"{sheet_name}_quarterly_data.csv", csv_content.getvalue())
                
                # Add summary data CSV files
                for summary_name, summary_records in all_summary_data.items():
                    if summary_records:
                        csv_content = io.StringIO()
                        headers = list(summary_records[0].keys()) if isinstance(summary_records[0], dict) else [f"col_{i}" for i in range(len(summary_records[0]))]
                        writer = csv.DictWriter(csv_content, fieldnames=headers)
                        writer.writeheader()
                        for record in summary_records:
                            writer.writerow(record if isinstance(record, dict) else dict(zip(headers, record)))
                        
                        zip_file.writestr(f"{summary_name}.csv", csv_content.getvalue())
            
            # Return the zip file
            def zip_generator():
                with open(zip_path, 'rb') as f:
                    while True:
                        chunk = f.read(8192)
                        if not chunk:
                            break
                        yield chunk
            
            return StreamingResponse(
                zip_generator(),
                media_type="application/zip",
                headers={
                    "Content-Disposition": f"attachment; filename=quarterly_export_Q{'-'.join(map(str, quarter_list))}_{'-'.join(map(str, year_list))}.zip"
                }
            )
        
        elif format == "xlsx":
            try:
                import pandas as pd
                import tempfile
                
                # Create temporary Excel file
                temp_dir = tempfile.mkdtemp()
                excel_path = f"{temp_dir}/quarterly_export.xlsx"
                
                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    # Add quarterly data sheets
                    for sheet_name, records in all_quarterly_data.items():
                        if records:
                            df = pd.DataFrame(records)
                            # Truncate sheet name if too long (Excel limit is 31 characters)
                            safe_sheet_name = sheet_name[:28] + "..." if len(sheet_name) > 31 else sheet_name
                            df.to_excel(writer, sheet_name=safe_sheet_name, index=False)
                    
                    # Add summary data sheets
                    for summary_name, summary_records in all_summary_data.items():
                        if summary_records:
                            df = pd.DataFrame(summary_records)
                            # Truncate sheet name if too long
                            safe_sheet_name = summary_name[:28] + "..." if len(summary_name) > 31 else summary_name
                            df.to_excel(writer, sheet_name=safe_sheet_name, index=False)
                
                # Return the Excel file
                def excel_generator():
                    with open(excel_path, 'rb') as f:
                        while True:
                            chunk = f.read(8192)
                            if not chunk:
                                break
                            yield chunk
                
                return StreamingResponse(
                    excel_generator(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={
                        "Content-Disposition": f"attachment; filename=quarterly_export_Q{'-'.join(map(str, quarter_list))}_{'-'.join(map(str, year_list))}.xlsx"
                    }
                )
                
            except ImportError as e:
                logger.warning(f"Pandas/openpyxl not available for XLSX export: {str(e)}")
                # Fallback to simple XLSX using openpyxl directly
                try:
                    from openpyxl import Workbook
                    import tempfile
                    
                    # Create temporary Excel file using openpyxl directly
                    temp_dir = tempfile.mkdtemp()
                    excel_path = f"{temp_dir}/quarterly_export.xlsx"
                    
                    wb = Workbook()
                    # Remove default sheet
                    wb.remove(wb.active)
                    
                    # Add quarterly data sheets
                    for sheet_name, records in all_quarterly_data.items():
                        if records:
                            safe_sheet_name = sheet_name[:28] + "..." if len(sheet_name) > 31 else sheet_name
                            ws = wb.create_sheet(title=safe_sheet_name)
                            
                            # Add headers
                            headers = list(records[0].keys()) if records else []
                            for col, header in enumerate(headers, 1):
                                ws.cell(row=1, column=col, value=header)
                            
                            # Add data
                            for row_idx, record in enumerate(records, 2):
                                for col_idx, header in enumerate(headers, 1):
                                    ws.cell(row=row_idx, column=col_idx, value=str(record.get(header, "")))
                    
                    # Add summary data sheets
                    for summary_name, summary_records in all_summary_data.items():
                        if summary_records:
                            safe_sheet_name = summary_name[:28] + "..." if len(summary_name) > 31 else summary_name
                            ws = wb.create_sheet(title=safe_sheet_name)
                            
                            # Add headers
                            headers = list(summary_records[0].keys()) if summary_records else []
                            for col, header in enumerate(headers, 1):
                                ws.cell(row=1, column=col, value=header)
                            
                            # Add data
                            for row_idx, record in enumerate(summary_records, 2):
                                for col_idx, header in enumerate(headers, 1):
                                    ws.cell(row=row_idx, column=col_idx, value=str(record.get(header, "")))
                    
                    # Save workbook
                    wb.save(excel_path)
                    
                    # Return the Excel file
                    def excel_generator():
                        with open(excel_path, 'rb') as f:
                            while True:
                                chunk = f.read(8192)
                                if not chunk:
                                    break
                                yield chunk
                    
                    return StreamingResponse(
                        excel_generator(),
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={
                            "Content-Disposition": f"attachment; filename=quarterly_export_Q{'-'.join(map(str, quarter_list))}_{'-'.join(map(str, year_list))}.xlsx"
                        }
                    )
                    
                except ImportError:
                    raise HTTPException(
                        status_code=status.HTTP_501_NOT_IMPLEMENTED,
                        detail="XLSX export requires pandas and openpyxl packages. Please use CSV or JSON format instead."
                    )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in export_quarterly_sheet_data: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to export quarterly sheet data: {str(e)}"
        )


@router.get("/agent-mis/{agent_code}")
async def get_agent_mis_data(
    agent_code: str,
    quarter: int = Query(..., ge=1, le=4, description="Quarter number (1-4)"),
    year: int = Query(..., ge=2020, le=2030, description="Year"),
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(50, ge=1, le=500, description="Items per page"),
    current_user = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    _rbac_check = Depends(require_admin_read)
):
    """
    Get complete quarterly MIS data for a specific agent including summary
    
    **Admin/SuperAdmin only endpoint**
    
    Returns quarterly sheet data filtered for the specified agent with:
    - Complete quarterly sheet data (all fields) for the agent including new fields
    - Both MATCH = TRUE and MATCH = FALSE records
    - Agent's summary sheet data with calculations
    - All quarterly sheet fields included (70+ fields)
    
    **Parameters:**
    - **agent_code**: Agent code to filter data for
    - **quarter**: Quarter number (1-4) for the quarterly sheet
    - **year**: Year for the quarterly sheet
    
    **Returns:**
    - Complete quarterly sheet records for the agent
    - Agent's summary data from Summary sheet
    - Statistics and calculations
    - All fields from quarterly sheets including new fields:
      * "Agent Total PO Amount": Total agent policy office amount
      * "Actual Agent_PO%": Actual agent policy office percentage
    - All other 70+ quarterly sheet fields (not filtered)
    """
    try:
        logger.info(f"Fetching quarterly MIS data for agent: {agent_code}, Q{quarter}-{year}")
        
        # Get quarterly sheet data for the agent
        quarterly_result = await mis_helpers.get_quarterly_sheet_agent_data(
            agent_code=agent_code,
            quarter=quarter,
            year=year,
            page=page,
            page_size=page_size
        )
        
        # Get summary sheet data for the agent
        summary_result = await mis_helpers.get_agent_summary_data(agent_code=agent_code)
        
        if not quarterly_result:
            logger.warning(f"No quarterly data found for agent: {agent_code} in Q{quarter}-{year}")
            quarterly_result = {
                "records": [],
                "total_count": 0,
                "page": page,
                "page_size": page_size,
                "total_pages": 0
            }
        
        if not summary_result:
            logger.warning(f"No summary data found for agent: {agent_code}")
            summary_result = {}
        
        return {
            "agent_code": agent_code,
            "quarter": quarter,
            "year": year,
            "quarterly_data": {
                "records": quarterly_result["records"],
                "total_count": quarterly_result["total_count"],
                "page": quarterly_result["page"],
                "page_size": quarterly_result["page_size"],
                "total_pages": quarterly_result["total_pages"]
            },
            "summary_data": summary_result,
            "sheet_name": f"Q{quarter}-{year}"
        }
        
    except Exception as e:
        logger.error(f"Error in get_agent_mis_data: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch agent MIS data for agent {agent_code} in Q{quarter}-{year}"
        )


@router.get("/my-mis", response_model=AgentMISResponse)
async def get_my_mis_data(
    quarter: int = Query(..., ge=1, le=4, description="Quarter number (1-4)"),
    year: int = Query(..., ge=2020, le=2030, description="Year"),
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(50, ge=1, le=500, description="Items per page"),
    current_user = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    _rbac_check = Depends(require_permission("policies", "read"))
):
    """
    Get agent's own quarterly MIS data with summary
    
    **Agent endpoint**
    
    Returns the authenticated agent's own quarterly sheet data with:
    - Data from specified quarterly sheet (Q{quarter}-{year})
    - Only records where MATCH = TRUE from quarterly sheet
    - Agent's summary data from Summary sheet (MATCH = TRUE only)
    - Limited filtered fields as per AgentMISRecord schema (only essential fields)
    - Calculated statistics (number of policies, running balance, total net premium)
    - Agent can only see their own data based on their user profile
    
    **Parameters:**
    - **quarter**: Quarter number (1-4) for the quarterly sheet
    - **year**: Year for the quarterly sheet
    - **page**: Page number for pagination
    - **page_size**: Items per page
    
    **Security:**
    - Uses authenticated user's ID to lookup agent code from UserProfile
    - Agent cannot access other agents' data
    - Filtered data excludes sensitive broker information
    
    **Returns:**
    - Quarterly sheet records for the agent (filtered fields)
    - Agent's summary data from Summary sheet
    - Statistics and pagination info
    
    Fields included in AgentMISRecord (limited set as requested):
    - Booking date: Policy booking date
    - Policy start date: Policy start date
    - Policy end date: Policy end date  
    - Policy number: Policy number
    - Insurer name: Insurance company name
    - Broker name: Broker name
    - Gross premium: Gross premium amount
    - Net premium: Net premium amount
    - Commissionable premium: Commissionable premium amount
    - Agent Total PO Amount: Total agent policy office amount
    - Actual Agent PO%: Actual agent policy office percentage
    
    **New Fields Support:**
    - "Agent Total PO Amount" is now included as agent_total_po_amount field
    - "Actual Agent_PO%" is now included as actual_agent_po_percent field
    """
    try:
        user_id = current_user["user_id"]
        
        # Get the agent's code from UserProfile table
        from sqlalchemy import select
        from models import UserProfile
        
        query = select(UserProfile.agent_code).where(UserProfile.user_id == user_id)
        result = await db.execute(query)
        agent_code = result.scalar_one_or_none()
        
        if not agent_code:
            logger.warning(f"No agent code found for user: {user_id}")
            return AgentMISResponse(
                records=[],
                stats=AgentMISStats(
                    number_of_policies=0,
                    running_balance=0.0,
                    total_net_premium=0.0
                ),
                total_count=0,
                page=page,
                page_size=page_size,
                total_pages=0
            )
        
        logger.info(f"Fetching quarterly MIS data for agent: {agent_code} (user: {user_id}) from Q{quarter}-{year}")
        
        # Get quarterly sheet data for the agent (MATCH = TRUE only)
        quarterly_result = await mis_helpers.get_quarterly_sheet_agent_filtered_data(
            agent_code=agent_code,
            quarter=quarter,
            year=year,
            page=page,
            page_size=page_size,
            match_only=True  # Only MATCH = TRUE records
        )
        
        if not quarterly_result:
            logger.warning(f"No quarterly MIS data found for agent: {agent_code} in Q{quarter}-{year}")
            return AgentMISResponse(
                records=[],
                stats=AgentMISStats(
                    number_of_policies=0,
                    running_balance=0.0,
                    total_net_premium=0.0
                ),
                total_count=0,
                page=page,
                page_size=page_size,
                total_pages=0
            )
            
        # Convert records to AgentMISRecord objects with field mapping
        agent_records = []
        for record_dict in quarterly_result["records"]:
            # Helper function to safely convert to string
            def safe_str(value):
                return str(value).strip() if value is not None and str(value).strip() else None
            
            # Map quarterly sheet fields to AgentMISRecord fields - only specific fields requested
            mapped_record = {
                "booking_date": safe_str(record_dict.get("Booking Date(Click to select Date)") or record_dict.get("booking_date")),
                "policy_start_date": safe_str(record_dict.get("Policy Start Date") or record_dict.get("policy_start_date")),
                "policy_end_date": safe_str(record_dict.get("Policy End Date") or record_dict.get("policy_end_date")),
                "policy_number": safe_str(record_dict.get("Policy number") or record_dict.get("policy_number")),
                "insurer_name": safe_str(record_dict.get("Insurer name") or record_dict.get("insurer_name")),
                "broker_name": safe_str(record_dict.get("Broker Name") or record_dict.get("broker_name")),
                "gross_premium": safe_str(record_dict.get("Gross premium") or record_dict.get("gross_premium")),
                "net_premium": safe_str(record_dict.get("Net premium") or record_dict.get("net_premium")),
                "commissionable_premium": safe_str(record_dict.get("Commissionable Premium") or record_dict.get("commissionable_premium")),
                "agent_total_po_amount": safe_str(record_dict.get("Agent Total PO Amount") or record_dict.get("agent_total_po_amount")),
                "actual_agent_po_percent": safe_str(record_dict.get("Actual Agent_PO%") or record_dict.get("actual_agent_po_percent"))
            }
            agent_records.append(AgentMISRecord(**mapped_record))
            
        # Calculate stats from the quarterly data
        stats = AgentMISStats(
            number_of_policies=quarterly_result.get("total_count", 0),
            running_balance=quarterly_result.get("stats", {}).get("running_balance", 0.0),
            total_net_premium=quarterly_result.get("stats", {}).get("total_net_premium", 0.0),
            commissionable_premium=quarterly_result.get("stats", {}).get("commissionable_premium", 0.0)
        )
        
        return AgentMISResponse(
            records=agent_records,
            stats=stats,
            total_count=quarterly_result["total_count"],
            page=quarterly_result["page"],
            page_size=quarterly_result["page_size"],
            total_pages=quarterly_result["total_pages"]
        )
        
    except Exception as e:
        logger.error(f"Error in get_my_mis_data for user {user_id}, Q{quarter}-{year}: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch your MIS data for Q{quarter}-{year}"
        )


@router.get("/policy-documents", response_model=PolicyDocumentsResponse)
async def get_policy_documents(
    policy_number: str = Query(..., description="Policy number to search for (supports special characters like HDFC/TW/2025/CUT128)"),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_admin_read)
):
    """
    Get policy PDF URL and additional documents URLs for a specific policy number.
    
    Searches both the Policy and CutPay tables for document URLs associated with the policy number.
    Returns information from both tables if found, with preference given to Policy table data.
    
    **Parameters:**
    - **policy_number**: The policy number to search for (query parameter)
    
    **Returns:**
    - Policy PDF URL and additional documents URL
    - Source information indicating which table(s) contained the data
    - Boolean flags indicating presence in each table
    
    **Usage Example:**
    - GET /mis/policy-documents?policy_number=HDFC/TW/2025/CUT128
    
    **Access:** Requires admin read permissions
    """
    try:
        from models import Policy, CutPay
        
        # Search in Policy table
        policy_stmt = select(Policy.policy_pdf_url, Policy.additional_documents).where(
            Policy.policy_number == policy_number
        )
        policy_result = await db.execute(policy_stmt)
        policy_data = policy_result.first()
        
        # Search in CutPay table
        cutpay_stmt = select(CutPay.policy_pdf_url, CutPay.additional_documents).where(
            CutPay.policy_number == policy_number
        )
        cutpay_result = await db.execute(cutpay_stmt)
        cutpay_data = cutpay_result.first()
        
        # Determine response data
        found_in_policy = policy_data is not None
        found_in_cutpay = cutpay_data is not None
        
        if not found_in_policy and not found_in_cutpay:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Policy number '{policy_number}' not found in Policy or CutPay tables"
            )
        
        # Prefer Policy table data if available, otherwise use CutPay data
        if found_in_policy:
            policy_pdf_url = policy_data.policy_pdf_url
            additional_documents = policy_data.additional_documents
            source = "policy"
        else:
            policy_pdf_url = cutpay_data.policy_pdf_url
            additional_documents = cutpay_data.additional_documents
            source = "cutpay"
        
        return PolicyDocumentsResponse(
            policy_number=policy_number,
            policy_pdf_url=policy_pdf_url,
            additional_documents=additional_documents,
            source=source,
            found_in_policy_table=found_in_policy,
            found_in_cutpay_table=found_in_cutpay
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching policy documents for policy {policy_number}: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch policy documents for policy number: {policy_number}"
        )
