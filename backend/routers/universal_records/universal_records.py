"""
Universal Records Router

This router handles universal record uploads with insurer-specific mapping
for data reconciliation and system synchronization.
"""

import io
import csv
import logging
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, status, File, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from config import get_db
from dependencies.rbac import require_admin_read, require_admin_write, get_current_user
from . import helpers
from .schemas import (
    UniversalRecordUploadResponse,
    AvailableInsurersResponse,
    InsurerSelectionRequest,
    CSVPreviewRequest,
    CSVPreviewResponse,
    ReconciliationSummaryResponse
)

router = APIRouter(prefix="/universal-records", tags=["Universal Records"])
logger = logging.getLogger(__name__)


@router.get("/insurers", response_model=AvailableInsurersResponse)
async def get_available_insurers(
    current_user = Depends(get_current_user),
    _rbac_check = Depends(require_admin_read)
):
    """
    Get list of available insurer mappings
    
    **Admin only endpoint**
    
    Returns all available insurance companies that have configured
    header mappings for universal record processing.
    """
    
    try:
        insurers = helpers.get_available_insurers()
        
        return AvailableInsurersResponse(
            insurers=insurers,
            total_count=len(insurers)
        )
        
    except Exception as e:
        logger.error(f"Error fetching available insurers: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch available insurers"
        )


@router.post("/preview", response_model=CSVPreviewResponse)
async def preview_csv_with_insurer_mapping(
    file: UploadFile = File(..., description="Universal record CSV file"),
    insurer_name: str = None,
    preview_rows: int = 5,
    current_user = Depends(get_current_user),
    _rbac_check = Depends(require_admin_read)
):
    """
    Preview CSV file with selected insurer mapping applied
    
    **Admin only endpoint**
    
    This endpoint allows admins to preview how a CSV file will be processed
    with a specific insurer mapping before actual upload. Shows:
    
    - Original headers from CSV
    - Mapped headers after applying insurer configuration  
    - Sample data rows with mapping applied
    - Any unmapped headers that will be ignored
    - Total row count in the file
    
    **Parameters:**
    - `file`: CSV file to preview
    - `insurer_name`: Name of insurer mapping to apply
    - `preview_rows`: Number of sample rows to show (default: 5)
    """
    
    try:
        if not file.filename.lower().endswith('.csv'):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Only CSV files are allowed"
            )
        
        if not insurer_name:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="insurer_name parameter is required"
            )

        file_content = await file.read()
        csv_content = file_content.decode('utf-8')
        
        if not csv_content.strip():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="CSV file is empty"
            )
        
        preview_response = helpers.preview_csv_with_mapping(
            csv_content=csv_content,
            insurer_name=insurer_name,
            preview_rows=preview_rows
        )
        
        logger.info(f"CSV preview generated for {insurer_name} by admin {current_user['user_id']}")
        
        return preview_response
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error previewing CSV: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to preview CSV: {str(e)}"
        )


@router.post("/upload", response_model=UniversalRecordUploadResponse)
async def upload_universal_record(
    file: UploadFile = File(..., description="Universal record CSV file"),
    insurer_name: str = None,
    current_user = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    _rbac_check = Depends(require_admin_write)
):
    """
    Upload and process universal record CSV with insurer-specific mapping
    
    **Admin only endpoint**
    
    This endpoint processes a universal record CSV file from insurance companies
    using insurer-specific header mappings. The universal record is considered 
    the source of truth and will be used to:
    
    1. **Update existing records** in Google master sheet where data mismatches are found
    2. **Add missing records** that exist in universal record but not in master sheet
    3. **Set MATCH STATUS to TRUE** for all updated/added records
    4. **Generate detailed report** of all changes made
    
    **Process Flow:**
    1. Apply insurer-specific header mapping to CSV data
    2. Parse and validate mapped content
    3. Get existing records from Google master sheet for the insurer
    4. For each record in uploaded sheet:
       - Use mapping to align headers with master sheet format
       - Compare with master sheet records (by policy number)
       - If present, update mapped fields and set MATCH STATUS to TRUE
       - If not present, add to master sheet and set MATCH STATUS to TRUE
    5. Generate comprehensive reconciliation report
    
    **Parameters:**
    - `file`: CSV file containing universal records
    - `insurer_name`: Name of insurer mapping to use (required)
    
    **Returns:**
    - Detailed report showing what was updated/added in master sheet
    - Processing statistics and timing
    - List of any errors encountered
    - Insurer-specific mapping information
    """
    
    try:
        if not file.filename.lower().endswith('.csv'):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Only CSV files are allowed"
            )
        
        if not insurer_name:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="insurer_name parameter is required"
            )

        file_content = await file.read()
        csv_content = file_content.decode('utf-8')
        
        if not csv_content.strip():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="CSV file is empty"
            )
        
        admin_user_id = current_user["user_id"]

        # Process universal record with insurer mapping
        report = await helpers.process_universal_record_csv(
            db=db,
            csv_content=csv_content,
            insurer_name=insurer_name,
            admin_user_id=admin_user_id
        )
        
        logger.info(f"Universal record processed by admin {admin_user_id}")
        logger.info(f"Insurer: {insurer_name}, File: {file.filename}, Size: {len(file_content)} bytes")
        logger.info(f"Processed {report.stats.total_records_processed} records in {report.stats.processing_time_seconds:.2f} seconds")
        
        return UniversalRecordUploadResponse(
            message=f"Universal record processed successfully for {insurer_name}. "
                   f"Processed {report.stats.total_records_processed} records, "
                   f"updated {report.stats.total_records_updated}, "
                   f"added {report.stats.total_records_added} in "
                   f"{report.stats.processing_time_seconds:.2f} seconds. "
                   f"All records marked with MATCH STATUS = TRUE.",
            report=report,
            processing_time_seconds=report.stats.processing_time_seconds
        )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error processing universal record: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to process universal record: {str(e)}"
        )


@router.get("/template")
async def download_universal_record_template(
    insurer_name: Optional[str] = None,
    current_user = Depends(get_current_user),
    _rbac_check = Depends(require_admin_read)
):
    """
    Download XLSX template for universal record upload
    
    **Admin only endpoint**
    
    Returns an Excel template file showing the expected format for universal record uploads.
    If an insurer name is provided, the template will use that insurer's specific headers.
    Otherwise, it returns a generic template with master sheet headers.
    
    **Parameters:**
    - `insurer_name`: Optional insurer name to get insurer-specific template
    
    **Returns:**
    - XLSX file with appropriate headers and sample data
    """
    
    try:
        if insurer_name:
            # Get insurer-specific template
            insurer_mapping = helpers.get_insurer_mapping(insurer_name)
            if not insurer_mapping:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"No mapping found for insurer: {insurer_name}"
                )
            
            # Use insurer's original headers
            headers = list(insurer_mapping.keys())
            filename = f"universal_record_template_{insurer_name.replace(' ', '_').lower()}.xlsx"
            
            # Sample data matching insurer headers
            sample_data = []
            for header in headers:
                if "Policy" in header or "Number" in header:
                    sample_data.append("POL-2024-001")
                elif "Product" in header or "Type" in header:
                    sample_data.append("Motor Insurance")
                elif "Agent" in header and "Code" in header:
                    sample_data.append("AGT001")
                elif "Premium" in header and "Gross" in header:
                    sample_data.append("25000.00")
                elif "Date" in header:
                    sample_data.append("2024-01-01")
                else:
                    sample_data.append("Sample Value")
        
        else:
            # Generic template with master sheet headers
            headers = [
                'policy_number',
                'policy_type',
                'insurance_type', 
                'agent_code',
                'broker_name',
                'insurance_company',
                'vehicle_type',
                'registration_number',
                'vehicle_class',
                'vehicle_segment',
                'gross_premium',
                'gst',
                'net_premium',
                'od_premium',
                'tp_premium',
                'start_date',
                'end_date',
                'cut_pay_amount',
                'commission_grid',
                'agent_commission_given_percent',
                'payment_by',
                'amount_received',
                'payment_method',
                'payment_source',
                'transaction_date',
                'payment_date',
                'notes'
            ]
            filename = "universal_record_template_generic.xlsx"
            
            sample_data = [
                'POL-2024-001',  # policy_number
                'Motor Insurance',  # policy_type
                'Comprehensive',  # insurance_type
                'AGT001',  # agent_code
                'ABC Insurance Brokers',  # broker_name
                'ICICI Lombard',  # insurance_company
                'Car',  # vehicle_type
                'MH12AB1234',  # registration_number
                'Private Car',  # vehicle_class
                'Sedan',  # vehicle_segment
                25000.00,  # gross_premium
                4500.00,  # gst
                20500.00,  # net_premium
                15000.00,  # od_premium
                5500.00,  # tp_premium
                '2024-01-01',  # start_date
                '2024-12-31',  # end_date
                2500.00,  # cut_pay_amount
                '10%',  # commission_grid
                12.5,  # agent_commission_given_percent
                'John Smith',  # payment_by
                2500.00,  # amount_received
                'Bank Transfer',  # payment_method
                'Company Account',  # payment_source
                '2024-01-15',  # transaction_date
                '2024-01-20',  # payment_date
                'Sample transaction'  # notes
            ]
        
        # Create XLSX template using openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Template"
            
            # Add headers
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}1'] = header
                ws[f'{col_letter}1'].font = ws[f'{col_letter}1'].font.copy(bold=True)
            
            # Add sample data
            for col_num, value in enumerate(sample_data, 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}2'] = value
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO
            excel_output = io.BytesIO()
            wb.save(excel_output)
            excel_output.seek(0)
            
            return StreamingResponse(
                excel_output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
            
        except ImportError:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="XLSX format not supported. openpyxl package not installed."
            )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating template: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to generate template"
        )


@router.get("/reconciliation/summary", response_model=ReconciliationSummaryResponse)
async def get_reconciliation_summary(
    insurer_name: Optional[str] = None,
    current_user = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    _rbac_check = Depends(require_admin_read)
):
    """
    Get reconciliation summary showing system vs universal record comparison
    
    **Admin only endpoint**
    
    Provides a high-level summary of reconciliation status, including:
    - Total records in system vs universal records
    - Match/mismatch statistics
    - Data variance percentages
    - Coverage analysis
    - Top fields with mismatches
    
    **Parameters:**
    - `insurer_name`: Optional filter by specific insurer
    
    **Returns:**
    - Comprehensive reconciliation statistics
    - Variance and coverage percentages
    - Mismatch analysis by field
    """
    
    try:
        summary = await helpers.generate_reconciliation_summary(
            db=db,
            insurer_name=insurer_name
        )
        
        logger.info(f"Reconciliation summary generated by admin {current_user['user_id']}")
        if insurer_name:
            logger.info(f"Filtered by insurer: {insurer_name}")
        
        return summary
        
    except Exception as e:
        logger.error(f"Error generating reconciliation summary: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to generate reconciliation summary"
        )


@router.get("/mappings/{insurer_name}")
async def get_insurer_mapping(
    insurer_name: str,
    current_user = Depends(get_current_user),
    _rbac_check = Depends(require_admin_read)
):
    """
    Get header mapping configuration for specific insurer
    
    **Admin only endpoint**
    
    Returns the header mapping configuration used to transform
    insurer-specific CSV headers to master sheet format.
    
    **Parameters:**
    - `insurer_name`: Name of the insurer
    
    **Returns:**
    - Mapping configuration showing insurer headers -> master headers
    """
    
    try:
        mapping = helpers.get_insurer_mapping(insurer_name)
        
        if not mapping:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"No mapping found for insurer: {insurer_name}"
            )
        
        return {
            "insurer_name": insurer_name,
            "header_mappings": mapping,
            "total_mappings": len(mapping)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching insurer mapping: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch insurer mapping"
        )
